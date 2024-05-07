import random
import pandas as pd
import xlsxwriter
import xlsxwriter.format
import openpyxl as opxl
import os
from tkinter import messagebox

def acumular_probabilidades(probabilidades):
    acumulados = [probabilidades[0]]
    for i in range(1, len(probabilidades)):
        acumulados.append(probabilidades[i] + acumulados[i-1])
    return acumulados


def clasificar_numero_aleatorio(rnd, clases, probabilidad_x_clase):
    vector_probabilidades_acumuladas = acumular_probabilidades(probabilidad_x_clase)
    for i, prob in enumerate(vector_probabilidades_acumuladas):
        if rnd < prob:
            return clases[i]
        


def generar_numeros_aleatorios(n=1, generar_nuevos=True):
    tipo_generacion = ('puerta', 'género', 'venta', 'suscripciones')
    
    if generar_nuevos:
        v_puerta = [random.random() for _ in range(n)]
        v_genero = [random.random() for _ in range(n)]
        v_venta = [random.random() for _ in range(n)]
        v_suscripciones = [random.random() for _ in range(n)]

        vectores_aleatorios = [v_puerta, v_genero, v_venta, v_suscripciones]
        
        for i in range(len(vectores_aleatorios)):
            with open(f"csvs/numeros_aleatorios_{tipo_generacion[i]}.csv", "wt") as f:
                vec_to_join = [str(valor) for valor in vectores_aleatorios[i]]
                f.write( "\n".join(vec_to_join) )
    
    else:
        
        vectores_aleatorios = []
        for i in range(4):
            with open(f"csvs/numeros_aleatorios_{tipo_generacion[i]}.csv", "rt") as f:
                vec = [float(linea.strip()) for linea in f.readlines()]
                vectores_aleatorios.append(vec)
        
    return vectores_aleatorios


def validar_i_j(i, j, n):

    try:
        i = int(i)
        j = int(j)
    except ValueError:
        return False
    
    if j > n:
        return False
    if j < 1:
        return False
    if i < 1:
        return False
    if i + j > n + 1:
        return False

    return i, j



def validar_n(n_visitas):

    try:
        n_visitas = int(n_visitas)
    except ValueError:
        return False
    
    if not 10 <= n_visitas <= 1000000:
        return False
    
    return n_visitas



def validar_parametros(puerta, genero, venta_sra, venta_sr, utilidad):

    probs = []

    for i, prob in enumerate((puerta, genero, venta_sra, venta_sr, utilidad)):
        

        try:
            probabilidad = float(prob)
        except ValueError:
            # messagebox.showerror("Error", "Por favor ingrese valores numéricos válidos para todos los parámetros.")
            return False

        # Validar rangos y formatos
        if i != 4 and not 0 <= probabilidad < 1:
            # messagebox.showerror("Error", "Las probabilidades deben estar entre 0 y 1.")
            return False
        
         # Validar utilidad positiva
        if i == 4 and probabilidad <= 0:
            # messagebox.showerror("Error", "La utilidad debe ser positiva.")
            return False
        
        probs.append(probabilidad)

    return probs



# Hay que ver si se pueden pasar asi los frame y desglozarlos
def validar_distribuciones(dist_sra, dist_sr):
    
    try:
        # Obtener los valores de las entradas en el frame de distribución de suscripciones para señoras
        valores_sra = [float(entry.get()) for entry in dist_sra]
        valores_sra = [val for val in valores_sra if 0 <= val < 1]
    except ValueError:
        # Alguno de los parametros no es float
        return False 

    # Sumar los valores y verificar si la suma es igual a 1
    suma_sra = round(sum(valores_sra), 4)
    print(suma_sra)
    if suma_sra != 1 or len(valores_sra) != 3:
        # messagebox.showerror("Error", "La suma de las probabilidades para señoras debe ser igual a 1.")
        return False


    try:
        # Obtener los valores de las entradas en el frame de distribución de suscripciones para señores
        valores_sr = [float(entry.get()) for entry in dist_sr]
        valores_sr = [val for val in valores_sr if 0 <= val < 1]
    except ValueError:
        # Alguno de los parametros no es float
        return False

    # Sumar los valores y verificar si la suma es igual a 1
    suma_sr = round(sum(valores_sr), 4)
    if suma_sr != 1 or len(valores_sr) != 4:
        # messagebox.showerror("Error", "La suma de las probabilidades para señores debe ser igual a 1.")
        return False

    # Si todas las sumas son igual a 1, retornar True
    return (valores_sra, valores_sr)






def probabilidad_a_distribucion(probabilidad):
    probabilidad_distribuida = []
    probabilidad_distribuida.append(probabilidad)
    probabilidad_distribuida.append(1-probabilidad)

    return probabilidad_distribuida




























def generate_table(vector_estado, i, j, filepath="Tabla de simulacion.xlsx", auto_open=True):

    # Creamos el writer para escribir sobre el excel
    writer = pd.ExcelWriter(filepath, engine="xlsxwriter") 

    # Creamos el dataframe de pandas para escribirlo luego en el excel
    data_frame = pd.DataFrame(vector_estado[j-1:i+j-1])
    data_frame.columns = ['Iteracion', 'Rnd_Puerta', 'Puerta', "Rnd_Genero", 'Genero',
                          "Rnd_Venta", 'Venta', "Rnd_Suscripciones", 
                          'Suscripciones', 'Utilidad_venta', 'Total_utilidad', 
                          'Total_ventas', 'Probabilidad_venta']
    
    
    data_frame.to_excel(writer, sheet_name="Vector Estado", index=False)


    # print(data_frame)




    # Tomamos la hoja creada con pandas y formateamos las columnas para que sea legible
    worksheet = writer.sheets["Vector Estado"]
    # cell_format = workbook.add_format({'num_format': '#.####'})
    # cell_format.set_  # Enable text wrapping in cells (optional)

    column_widths = {
    'Iteracion': 10,
    "Rnd_Puerta": 20,
    'Puerta': 15,
    "Rnd_Genero": 20,
    'Genero': 10,
    "Rnd_Venta": 20,
    'Venta': 12,
    "Rnd_Suscripciones": 20,
    'Suscripciones': 15,
    'Utilidad_venta': 15,
    'Total_utilidad': 15,
    'Total_ventas': 15,
    'Probabilidad_venta': 20,
    }

    

    # Loop through columns and set widths
    for col_idx, col_name in enumerate(data_frame.columns):
        worksheet.set_column(col_idx, col_idx, column_widths[col_name])
        # if pd.api.types.is_numeric_dtype(data_frame[col_name]):
        #     worksheet.set_cell_format(col_idx, col_idx, cell_format)






    

    # Se cierra el writer
    writer.close()




# generate_table([[1, 2, 3, 4, 5, 6, 7, 8, 9],
#                 [2,22,33,44,55,66,77,88,99],
#                 [3,222,333,444,55,66,77,88,99],
#                 [4,22,42133,44,55,66,77,88,99],
#                 [5,22,33,44,55,66,77,88,99],
#                 [6,52522,33,44,55,66,77,88,99],
#                 [7,22,33,44,55,66,77,88,99444],], i=3, j=2)



def get_table(vector_estado, i, j, filepath="Tabla de simulacion.xlsx", auto_open=True):
    # Creamos el handler del workbook y la hoja
    wb = opxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Agregar primera fila de Headers
    ws.append(
        ['Iteracion', 'Rnd_Puerta', 'Puerta', "Rnd_Genero", 
         'Genero', "Rnd_Venta", 'Venta', "Rnd_Suscripciones", 
        'Suscripciones', 'Utilidad_venta', 'Total_utilidad', 
        'Total_ventas', 'Probabilidad_venta']
    )


    # Por cada fila a mostrar, redondear rnd's, y colocar -- en genero si es None
    for fila in vector_estado[j-1:j-1+i]:
        if fila[4] is None:
            fila[4] = "--"
        if fila[6] is None:
            fila[6] = "--"
        fila[1] = round(fila[1],4)
        fila[3] = round(fila[3],4)
        fila[5] = round(fila[5], 4)
        fila[7] = round(fila[7], 4)
        fila[-1] = round(fila[-1], 4)

        ws.append(fila)



    
    # Por cada columna de la tabla
    for letra in "ABCDEFGHIJKLM":
        # Ajustar ancho de columnas
        if ws.column_dimensions[letra].has_style:
            ws.column_dimensions[letra] = None
        ws.column_dimensions[letra].width = 20
        
        for cell in ws[f"{letra}:{letra}"]:
            # Alinear celdas al centro
            cell.alignment = opxl.styles.Alignment(horizontal="center")

            # borders
            cell.border = opxl.styles.Border(
                        left=opxl.styles.Side(border_style="thin", color="000000"),
                        right=opxl.styles.Side(border_style="thin", color="000000"),
                        top=opxl.styles.Side(border_style="thin", color="000000"),
                        bottom=opxl.styles.Side(border_style="thin", color="000000")
                        )


        
        # Style the first row (Headers)
        ws[f"{letra}1"].font = opxl.styles.Font(bold=True)
        ws[f"{letra}1"].fill = opxl.styles.PatternFill(patternType="solid", fgColor="C4C2C1")
        



    

    # Guardar la ultima fila siempre

    # create new sheet and append the headers row
    last_row_sheet = wb.create_sheet("Ultima Iteracion")
    last_row_sheet.append(['Iteracion', 'Rnd_Puerta', 'Puerta', "Rnd_Genero", 
         'Genero', "Rnd_Venta", 'Venta', "Rnd_Suscripciones", 
        'Suscripciones', 'Utilidad_venta', 'Total_utilidad', 
        'Total_ventas', 'Probabilidad_venta'])
    

    # Redondear la ultima iteracion y appendearla al sheet
    ultima_iteracion = vector_estado[-1]
    if ultima_iteracion[4] is None:
        ultima_iteracion[4] = "--"
    if ultima_iteracion[6] is None:
        ultima_iteracion[6] = "--"
    ultima_iteracion[1] = round(ultima_iteracion[1], 4)
    ultima_iteracion[3] = round(ultima_iteracion[3], 4)
    ultima_iteracion[5] = round(ultima_iteracion[5], 4)
    ultima_iteracion[7] = round(ultima_iteracion[7], 4)
    ultima_iteracion[-1] = round(ultima_iteracion[-1], 4)

    last_row_sheet.append(ultima_iteracion)


    # Estilizar todas las columnas y celdas de la nueva hoja
    for letra in "ABCDEFGHIJKLM":
        # Ajustar ancho de columnas
        if last_row_sheet.column_dimensions[letra].has_style:
            last_row_sheet.column_dimensions[letra] = None
        last_row_sheet.column_dimensions[letra].width = 20
        
        for cell in last_row_sheet[f"{letra}:{letra}"]:
            # Alinear al centro
            cell.alignment = opxl.styles.Alignment(horizontal="center")

            # borders
            cell.border = opxl.styles.Border(
                        left=opxl.styles.Side(border_style="thin", color="000000"),
                        right=opxl.styles.Side(border_style="thin", color="000000"),
                        top=opxl.styles.Side(border_style="thin", color="000000"),
                        bottom=opxl.styles.Side(border_style="thin", color="000000")
                        )


        
        # Style the first row (Headers)
        last_row_sheet[f"{letra}1"].font = opxl.styles.Font(bold=True)
        last_row_sheet[f"{letra}1"].fill = opxl.styles.PatternFill(patternType="solid", fgColor="C4C2C1")



    # Auto abrir el excel 
    if auto_open:
        os.startfile(filepath)
    


    wb.save(filepath)

# get_table()




# ve = [ [86, 0.04254699674622253, True, 0.12030977550659272, 'M', 0.597514290508959, False, 0.546727638080471, 0, 0.0, 3000.0, 10, 0.11627906976744186]   ,
# [87, 0.3720037738178468, True, 0.718278374680727, 'M', 0.828994524747727, False, 0.8765108175639132, 0, 0.0, 3000.0, 10, 0.11494252873563218]     ,
# [88, 0.8545743816844371, False, 0.49449238979713084, None, 0.9765103509595314, False, 0.9167373278838388, 0, 0.0, 3000.0, 10, 0.11363636363636363],
# [89, 0.4885392888618113, True, 0.2185904995054534, 'M', 0.6143612576412, False, 0.9245327570678554, 0, 0.0, 3000.0, 10, 0.11235955056179775]      ,
# [90, 0.8018832190079382, False, 0.7489149048768589, None, 0.8255770547578429, False, 0.6450235550014718, 0, 0.0, 3000.0, 10, 0.1111111111111111]  ,
# [91, 0.3879385962299686, True, 0.29786612270633184, 'M', 0.7085313863796281, False, 0.3271434756096916, 0, 0.0, 3000.0, 10, 0.10989010989010989]  ,
# [92, 0.6603878555195062, True, 0.6575662766754612, 'M', 0.39875320371783307, False, 0.8306796760919832, 0, 0.0, 3000.0, 10, 0.10869565217391304]  ,
# [93, 0.904800125205876, False, 0.39713345323274907, None, 0.1818470119059724, False, 0.5443239475187139, 0, 0.0, 3000.0, 10, 0.10752688172043011] ,
# [94, 0.6294433405777714, True, 0.961529174109437, 'H', 0.7494663047078605, False, 0.6714550007534907, 0, 0.0, 3000.0, 10, 0.10638297872340426]    ,
# [95, 0.17945123779549754, True, 0.43204025279987335, 'M', 0.0732944031081314, True, 0.3556010716090261, 1, 200.0, 3200.0, 11, 0.11578947368421053],
# [96, 0.8205773312596649, False, 0.8267156820912499, None, 0.18797561105700367, False, 0.42564412903567395, 0, 0.0, 3200.0, 11, 0.11458333333333333],
# [97, 0.8296117716858282, False, 0.1391205111169057, None, 0.4846879634199913, False, 0.6149924583355354, 0, 0.0, 3200.0, 11, 0.1134020618556701],
# [98, 0.2541347820247667, True, 0.898956352837531, 'H', 0.15102608082555347, True, 0.7786189763976838, 3, 600.0, 3800.0, 12, 0.12244897959183673],
# [99, 0.5319550784235477, True, 0.43431403583229533, 'M', 0.3754293008810856, False, 0.36913662218498655, 0, 0.0, 3800.0, 12, 0.12121212121212122],
# [100, 0.49264660582748176, True, 0.9625271307535179, 'H', 0.26804140059407067, False, 0.1592081544594558, 0, 0.0, 3800.0, 12, 0.12] ]




# get_table(ve, 10, 1, "Lacabra.xlsx")
import openpyxl as opxl
from openpyxl import load_workbook
import os


import shutil



def generate_table(vector_estado, ultima_fila, filepath="Tabla_de_simulacion.xlsx", auto_open=True):

    # Ruta del archivo fuente y del archivo de destino
    source_path = 'headers.xlsx'
    destination_path = 'tabla_cabra.xlsx'

    # Copiar el archivo
    shutil.copy(source_path, destination_path)


    # Crear el libro de trabajo y la hoja
    wb = load_workbook("tabla_cabra.xlsx")
    ws = wb.active
    ws.title = "Data"


    # Agregar datos a la tabla
    # for fila in vector_estado:
    #     try:
    #         # print("appendeo")
    #         ws.append(fila)
    #     except Exception:
    #         pass

    for i, fila in enumerate(vector_estado, start=3):
        for j, value in enumerate(fila, start=1):
            if isinstance(value, float):
                value = round(value, 4)
            value = str(value)
            ws.cell(row=i, column=j, value=value)

    wb.save("tabla_cabra.xlsx")

    if auto_open:
        os.startfile("tabla_cabra.xlsx")
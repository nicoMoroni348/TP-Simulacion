import pandas as pd
import openpyxl as opxl
import os
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

# def generate_table(vector_estado, i, j, filepath="Tabla_de_simulacion.xlsx", auto_open=True):

#     # Creamos el writer para escribir sobre el excel
#     writer = pd.ExcelWriter(filepath, engine="xlsxwriter") 

#     # Creamos el dataframe de pandas para escribirlo luego en el excel
#     data_frame = pd.DataFrame(vector_estado[j-1:i+j-1])
#     data_frame.columns = [
#         'F', 'Evento', 'Reloj', "Rnd_llegada_alumno", 
#          'tiempo_llegada_alumno', "Proxima_llegada_alumno", 'Rnd1_llegada_mantenimiento', "Rnd2_llegada_mantenimiento", 
#         'Proxima_llegada_mantenimiento', 'alumno_se_queda', 'Hora_regreso', 
#         'Cola_alumnos', 'Cola_mantenimiento',
#         'Rnd_fin_inscripcion', 'tiempo_fin_inscripcion', 'Rnd_fin_mantenimiento', 'tiempo_fin_mantenimiento',
#         'Equipo1_estado', 'Equipo1_Hora_fin', 'Equipo2_estado', 'Equipo2_Hora_fin',
#         'Equipo3_estado', 'Equipo3_Hora_fin', 'Equipo4_estado', 'Equipo4_Hora_fin',
#         'Equipo5_estado', 'Equipo5_Hora_fin', 'Equipo6_estado', 'Equipo6_Hora_fin',
#         'Contador_alumnos', 'Contador_alumnos_que_regresan',
#         'Porcentaje_alumnos_que_regresan', 'Porcentaje_alumnos_que_se_van', 'Contador_alumnos_atendidos', 'Acumulador_tiempo_espera', 'Promedio_tiempo_espera',
#         ]
    
    
#     data_frame.to_excel(writer, sheet_name="Vector Estado", index=False)


#     # print(data_frame)




#     # Tomamos la hoja creada con pandas y formateamos las columnas para que sea legible
#     worksheet = writer.sheets["Vector Estado"]
#     # cell_format = workbook.add_format({'num_format': '#.####'})
#     # cell_format.set_  # Enable text wrapping in cells (optional)

#     column_widths = {
#         'F': 10,
#         'Evento': 20,
#         'Reloj': 15,
#         'Rnd_llegada_alumno': 20,
#         'tiempo_llegada_alumno': 20,
#         'Proxima_llegada_alumno': 20,
#         'Rnd1_llegada_mantenimiento': 20,
#         'Rnd2_llegada_mantenimiento': 20,
#         'Proxima_llegada_mantenimiento': 20,
#         'alumno_se_queda': 15,
#         'Hora_regreso': 15,
#         'Cola_alumnos': 15,
#         'Cola_mantenimiento': 15,
#         'Rnd_fin_inscripcion': 20,
#         'tiempo_fin_inscripcion': 20,
#         'Rnd_fin_mantenimiento': 20,
#         'tiempo_fin_mantenimiento': 20,
#         'Equipo1_estado': 15,
#         'Equipo1_Hora_fin': 15,
#         'Equipo2_estado': 15,
#         'Equipo2_Hora_fin': 15,
#         'Equipo3_estado': 15,
#         'Equipo3_Hora_fin': 15,
#         'Equipo4_estado': 15,
#         'Equipo4_Hora_fin': 15,
#         'Equipo5_estado': 15,
#         'Equipo5_Hora_fin': 15,
#         'Equipo6_estado': 15,
#         'Equipo6_Hora_fin': 15,
#         'Contador_alumnos': 15,
#         'Contador_alumnos_que_regresan': 15,
#         'Porcentaje_alumnos_que_regresan': 15,
#         'Porcentaje_alumnos_que_se_van': 15,
#         'Contador_alumnos_atendidos': 15,
#         'Acumulador_tiempo_espera': 15,
#         'Promedio_tiempo_espera': 15,
#     }


#     # Loop through columns and set widths
#     for col_idx, col_name in enumerate(data_frame.columns):
#         worksheet.set_column(col_idx, col_idx, column_widths[col_name])
#         # if pd.api.types.is_numeric_dtype(data_frame[col_name]):
#         #     worksheet.set_cell_format(col_idx, col_idx, cell_format)

#     # Se cierra el writer
#     writer.close()

def generate_table(vector_estado, ultima_fila, filepath="Tabla_de_simulacion.xlsx", auto_open=True):

    # Crear el libro de trabajo y la hoja
    wb = opxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Agregar primera fila de Headers
    headers = [ 
                [ 'FILA', 'Evento', 'Reloj (minutos)', 'Llegada alumno', '', '', 'Llegada mantenimiento', '', '', '','Fin regreso alumno', '', 'Fin inscripción (i)', '', '', 'Fin mantenimiento (i)','', '', 'Variables Estadisticas', '', '', '', '', '', 'Colas', '', 'Mantenimiento', '', 'Equipo1', '','Equipo2', '','Equipo3', '','Equipo4', '', 'Equipo5', '','Equipo6', '','ALUMNOS'],
                 ['', '', '', 'RND', 'Tiempo', 'Próxima llegada', 'RND1', 'RND2', 'Tiempo', 'Próxima llegada','Se queda', 'Hora Regreso','RND', 'Tiempo', 'Próxima inscripcion', 'RND', 'Tiempo', 'Próximo Mantenimiento', 'Contador alumnos que llegan', 'Contador alumnos que regresan', 'Porcentaje alumnos que se van', 'Contador alumnos atendidos', 'Acumulador tiempos de espera', 'Promedios tiempos de espera', 'Cola alumnos', 'Cola mantenimiento', 'Estado', 'Maquinas restantes', 'Estado', 'HoraFin1', 'Estado', 'HoraFin2', 'Estado', 'HoraFin3', 'Estado', 'HoraFin4', 'Estado', 'HoraFin5', 'Estado', 'HoraFin6', 'IdAlumno', 'Estado','HoraLlegada', 'TiempoEspera', 'IdAlumno', 'Estado','HoraLlegada', 'TiempoEspera', 'IdAlumno', 'Estado','HoraLlegada', 'TiempoEspera', 'IdAlumno', 'Estado','HoraLlegada', 'TiempoEspera', 'IdAlumno', 'Estado','HoraLlegada', 'TiempoEspera']
            ]



    # Agregar headers al archivo
    for row in headers:
        ws.append(row)

    # Combinar celdas para los encabezados
    merge_ranges = [
        ('A1:A5'), ('B1:B5'), ('C1:C5'), ('D1:F1'),  ('G1:J1'), ('K1:L1'), ('M1:O1'), ('P1:R1'), ('S1:X1'), ('AA1:AB1'), ('Y1:Z1'), ('AA1:AB1'), ('AC1:AD1'), ('AE1:AF1'), ('AG1:AH1'), ('AI1:AJ1'), ('AK1:AL1'), ('AM1:AN1'), ('D2:D5'), ('E2:E5'), ('F2:F5'), ('G2:G5'), ('H2:H5'), ('I2:I5'), ('J2:J5'), ('K2:K5'), ('L2:L5'), ('M2:M5'),('N2:N5'), ('O2:O5'), ('P2:P5'), ('Q2:Q5'), ('R2:R5'), ('S2:S5'), ('T2:T5'), ('U2:U5'), ('V2:V5'), ('W2:W5'), ('X2:X5'), ('Y2:Y5'), ('Z2:Z5'), 
        ('AA2:AA5'), ('AB2:AB5'), ('AC2:AC5'), ('AD2:AD5'), ('AE2:AE5'), ('AF2:AF5'), ('AG2:AG5'), ('AH2:AH5'), ('AI2:AI5'), ('AJ2:AJ5'), ('AK2:AK5'), ('AL2:AL5'), ('AM2:AM5'), ('AN2:AN5'), ('AO2:AO5'), ('AP2:AP5'), ('AQ2:AQ5'), ('AR2:AR5'), ('AS2:AS5'), ('AT2:AT5'), ('AU2:AU5'), ('AV2:AV5'), ('AW2:AW5'), ('AX2:AX5'), ('AY2:AY5'), ('AZ2:AZ5'), ("BA2:BA5"),("BB2:BB5"), ("BC2:BC5"), ("BD2:BD5"), ("BE2:BE5"), ("BF2:BF5"), ("BG2:BG5"), ("BH2:BH5"), ("AO1:BH1")
    ]

    for merge_range in merge_ranges:
        ws.merge_cells(merge_range)

    # Formatear celdas
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True, wrap_text=True)
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell.font = Font(bold=True)
            cell.border = Border(
                        left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        top=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000")
                        )
            # Ajustar el tamaño de las celdas al contenido
            for column_cells in ws.columns:
                max_length = 0
                if isinstance(cell, opxl.cell.MergedCell):
                    continue
                column = cell.column_letter
                column = column_cells[0].column_letter  # Get the column name
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                # adjusted_width = (max_length + 2) * 1.2
                # adjusted_width = 50
                # ws.column_dimensions[column].width = adjusted_width

    abecedario = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")
    
    for letra in  abecedario + ["A"+ letra for letra in abecedario] + ["B" + letra for letra in abecedario]:
        # Ajustar ancho de columnas
        if ws.column_dimensions[letra].has_style:
            ws.column_dimensions[letra] = None
        ws.column_dimensions[letra].width = 20

    # # Agregar datos a la tabla
    # for fila in vector_estado:
    #     fila_formateada = []
    #     for idx, val in enumerate(fila):
    #         if isinstance(val, (int, float)):
    #             val = round(val, 4)
    #         if val is None:
    #             val = "--"
    #         fila_formateada.append(val)

        # fila_formateada = [str(el) for el in fila_formateada]
        # ws.append(fila_formateada)

    for i, fila in enumerate(vector_estado, start=6):
        for j, value in enumerate(fila, start=1):
            if isinstance(value, (int, float)):
                value = round(value, 4)
            if value is None:
                value = "--"
            value = str(value)
            ws.cell(row=i, column=j, value=value)



    # Ultima fila
    last_row_sheet: Worksheet = wb.create_sheet("Ultima Iteracion")

    for row in headers:
        last_row_sheet.append(row)

    for merge_range in merge_ranges:
        last_row_sheet.merge_cells(merge_range)

    
    for row in last_row_sheet.iter_rows(min_row=1,max_row=5):

        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True, wrap_text=True)
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell.font = Font(bold=True)
            cell.border = Border(
                        left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        top=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000")
                    )
        # Ajustar el tamaño de las celdas al contenido
        for column_cells in last_row_sheet.columns:
            max_length = 0
            if isinstance(cell, opxl.cell.MergedCell):
                continue
            # column = cell.column_letter
            # column = column_cells[0].column_letter  # Get the column name
            # # for cell in column_cells:
            # #     try:
            # #         if len(str(cell.value)) > max_length:
            # #             max_length = len(cell.value)
            # #     except:
            # #         pass
            # # adjusted_width = (max_length + 2) * 1.2
            # # last_row_sheet.column_dimensions[column].width = adjusted_width

    abecedario = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")
    
    for letra in  abecedario + ["A"+ letra for letra in abecedario] + ["B" + letra for letra in abecedario]:
        # Ajustar ancho de columnas
        if last_row_sheet.column_dimensions[letra].has_style:
            last_row_sheet.column_dimensions[letra] = None
        last_row_sheet.column_dimensions[letra].width = 20

    ult_fila_formateada = []
    for idx, val in enumerate(ultima_fila):
        
        if isinstance(val, (int, float)):
            val = round(val, 4)
        if val is None:
            val = "--"
        ult_fila_formateada.append(val)
    
    ult_fila_formateada = [str(el) for el in ult_fila_formateada]

    
    for j, value in enumerate(ult_fila_formateada, start=1):
        if isinstance(value, float):
            value = round(value, 4)
        value = str(value)
        last_row_sheet.cell(row=6, column=j, value=value)
    # print(ult_fila_formateada)

    # last_row_sheet.append(ult_fila_formateada)

    # Guardar el archivo
    wb.save(filepath)

    # Auto abrir el archivo
    if auto_open:
        os.startfile(filepath)


  # Por ejemplo, 10 filas desde el índice 1
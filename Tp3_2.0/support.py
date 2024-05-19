import openpyxl as opxl
import os


def validar_i_j(i, j, n):

    try:
        i = int(i)
        j = int(j)
    except ValueError:
        return False
    
    if j > n:
        return False
    if j < 1:
        return False
    if i < 1:
        return False
    if i + j > n + 1:
        return False

    return i, j


def validar_n(n_visitas):

    try:
        n_visitas = int(n_visitas)
    except ValueError:
        return False
    
    if not 10 <= n_visitas <= 1000000:
        return False
    
    return n_visitas


def validar_parametros(puerta, genero, venta_sra, venta_sr, utilidad):

    probs = []

    for i, prob in enumerate((puerta, genero, venta_sra, venta_sr, utilidad)):
        

        try:
            probabilidad = float(prob)
        except ValueError:
            # messagebox.showerror("Error", "Por favor ingrese valores numéricos válidos para todos los parámetros.")
            return False

        # Validar rangos y formatos
        if i != 4 and not 0 <= probabilidad < 1:
            # messagebox.showerror("Error", "Las probabilidades deben estar entre 0 y 1.")
            return False
        
         # Validar utilidad positiva
        if i == 4 and probabilidad <= 0:
            # messagebox.showerror("Error", "La utilidad debe ser positiva.")
            return False
        
        probs.append(probabilidad)

    return probs


# Hay que ver si se pueden pasar asi los frame y desglozarlos
def validar_distribuciones(dist_sra, dist_sr):
    
    try:
        # Obtener los valores de las entradas en el frame de distribución de suscripciones para señoras
        valores_sra = [float(entry.get()) for entry in dist_sra]
        valores_sra = [val for val in valores_sra if 0 <= val < 1]
    except ValueError:
        # Alguno de los parametros no es float
        return False 

    # Sumar los valores y verificar si la suma es igual a 1
    suma_sra = round(sum(valores_sra), 4)
    print(suma_sra)
    if suma_sra != 1 or len(valores_sra) != 3:
        # messagebox.showerror("Error", "La suma de las probabilidades para señoras debe ser igual a 1.")
        return False


    try:
        # Obtener los valores de las entradas en el frame de distribución de suscripciones para señores
        valores_sr = [float(entry.get()) for entry in dist_sr]
        valores_sr = [val for val in valores_sr if 0 <= val < 1]
    except ValueError:
        # Alguno de los parametros no es float
        return False

    # Sumar los valores y verificar si la suma es igual a 1
    suma_sr = round(sum(valores_sr), 4)
    if suma_sr != 1 or len(valores_sr) != 4:
        # messagebox.showerror("Error", "La suma de las probabilidades para señores debe ser igual a 1.")
        return False

    # Si todas las sumas son igual a 1, retornar True
    return (valores_sra, valores_sr)


def probabilidad_a_distribucion(probabilidad):
    probabilidad_distribuida = []
    probabilidad_distribuida.append(probabilidad)
    probabilidad_distribuida.append(1-probabilidad)

    return probabilidad_distribuida


def acumular_probabilidades(probabilidades):
    acumulados = [probabilidades[0]]
    for i in range(1, len(probabilidades)):
        acumulados.append(probabilidades[i] + acumulados[i-1])
    return acumulados


def clasificar_numero_aleatorio(rnd, clases, probabilidad_x_clase):
    vector_probabilidades_acumuladas = acumular_probabilidades(probabilidad_x_clase)
    for i, prob in enumerate(vector_probabilidades_acumuladas):
        if rnd < prob:
            return clases[i]
        

def get_table(vector_estado, ultima_fila, filepath="Tabla de simulacion.xlsx", auto_open=True):
    # Creamos el handler del workbook y la hoja
    wb = opxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Agregar primera fila de Headers
    ws.append(
        ['Iteracion', 'Rnd_Puerta', 'Puerta', "Rnd_Genero", 
         'Genero', "Rnd_Venta", 'Venta', "Rnd_Suscripciones", 
        'Suscripciones', 'Utilidad_venta', 'Total_utilidad', 
        'Total_ventas', 'Probabilidad_venta']
    )


    # Por cada fila a mostrar, redondear rnd's, y colocar -- en genero si es None
    for fila in vector_estado:
        if fila[4] is None:
            fila[4] = "--"
        if fila[6] is None:
            fila[6] = "--"
        fila[1] = round(fila[1],4) if fila[1] is not None else "--"
        fila[3] = round(fila[3],4) if fila[3] is not None else "--"
        fila[5] = round(fila[5], 4) if fila[5] is not None else "--"
        fila[7] = round(fila[7], 4) if fila[7] is not None else "--"
        fila[-1] = round(fila[-1], 4)

        ws.append(fila)


    # Por cada columna de la tabla
    for letra in "ABCDEFGHIJKLM":
        # Ajustar ancho de columnas
        if ws.column_dimensions[letra].has_style:
            ws.column_dimensions[letra] = None
        ws.column_dimensions[letra].width = 20
        
        for cell in ws[f"{letra}:{letra}"]:
            # Alinear celdas al centro
            cell.alignment = opxl.styles.Alignment(horizontal="center")

            # borders
            cell.border = opxl.styles.Border(
                        left=opxl.styles.Side(border_style="thin", color="000000"),
                        right=opxl.styles.Side(border_style="thin", color="000000"),
                        top=opxl.styles.Side(border_style="thin", color="000000"),
                        bottom=opxl.styles.Side(border_style="thin", color="000000")
                        )


        # Style the first row (Headers)
        ws[f"{letra}1"].font = opxl.styles.Font(bold=True)
        ws[f"{letra}1"].fill = opxl.styles.PatternFill(patternType="solid", fgColor="C4C2C1")
        

    # create new sheet and append the headers row
    last_row_sheet = wb.create_sheet("Ultima Iteracion")
    last_row_sheet.append(['Iteracion', 'Rnd_Puerta', 'Puerta', "Rnd_Genero", 
         'Genero', "Rnd_Venta", 'Venta', "Rnd_Suscripciones", 
        'Suscripciones', 'Utilidad_venta', 'Total_utilidad', 
        'Total_ventas', 'Probabilidad_venta'])
    

    # Redondear la ultima iteracion y appendearla al sheet
    if ultima_fila[4] is None:
        ultima_fila[4] = "--"
    if ultima_fila[6] is None:
        ultima_fila[6] = "--"
    ultima_fila[1] = round(ultima_fila[1], 4) if ultima_fila[1] is not None else "--"
    ultima_fila[3] = round(ultima_fila[3], 4) if ultima_fila[3] is not None else "--"
    ultima_fila[5] = round(ultima_fila[5], 4) if ultima_fila[5] is not None else "--"
    ultima_fila[7] = round(ultima_fila[7], 4) if ultima_fila[7] is not None else "--"
    ultima_fila[-1] = round(ultima_fila[-1], 4)

    last_row_sheet.append(ultima_fila)


    # Estilizar todas las columnas y celdas de la nueva hoja
    for letra in "ABCDEFGHIJKLM":
        # Ajustar ancho de columnas
        if last_row_sheet.column_dimensions[letra].has_style:
            last_row_sheet.column_dimensions[letra] = None
        last_row_sheet.column_dimensions[letra].width = 20
        
        for cell in last_row_sheet[f"{letra}:{letra}"]:
            # Alinear al centro
            cell.alignment = opxl.styles.Alignment(horizontal="center")

            # borders
            cell.border = opxl.styles.Border(
                        left=opxl.styles.Side(border_style="thin", color="000000"),
                        right=opxl.styles.Side(border_style="thin", color="000000"),
                        top=opxl.styles.Side(border_style="thin", color="000000"),
                        bottom=opxl.styles.Side(border_style="thin", color="000000")
                        )


        # Style the first row (Headers)
        last_row_sheet[f"{letra}1"].font = opxl.styles.Font(bold=True)
        last_row_sheet[f"{letra}1"].fill = opxl.styles.PatternFill(patternType="solid", fgColor="C4C2C1")


    # Auto abrir el excel 
    if auto_open:
        os.startfile(filepath)
    

    wb.save(filepath)